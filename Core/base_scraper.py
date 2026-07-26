import os
import csv
from abc import ABC, abstractmethod

class BaseScraper(ABC):
    def __init__(self, driver, target, output_dir="data"):
        """
        :param driver: An already authenticated Selenium WebDriver.
        :param target: The username, channel, or URL to scrape.
        """
        self.driver = driver
        self.target = target
        self.output_dir = output_dir
        
        # Ensure target is alphanumeric/clean for filenames
        safe_target = "".join(c for c in self.target if c.isalnum() or c in ('_', '-'))
        self.output_path = os.path.join(self.output_dir, f"{safe_target}_scraped_data.csv")

        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    @abstractmethod
    def navigate_to_page(self) -> bool:
        """
        Platform-specific logic to reach the target profile/channel.
        Must return True if successful, False otherwise.
        """
        pass

    @abstractmethod
    def scrape_all_posts(self, max_posts=None) -> list:
        """
        Platform-specific logic to scroll, extract, and return a list of dictionaries.
        """
        pass

    def _convert_persian_nums(self, text: str) -> str:
        """
        Universal utility for Iranian platforms to normalize digits.
        """
        persian_digits = "۰۱۲۳۴۵۶۷۸۹"
        arabic_digits = "٠١٢٣٤٥٦٧٨٩"
        english_digits = "0123456789"

        translation_table = str.maketrans(
            persian_digits + arabic_digits, english_digits * 2
        )
        return text.translate(translation_table)

    def save_to_csv(self, product_list: list) -> None:
        """
        Universal save method. Dynamically extracts headers from the first dictionary.
        Writes both a CSV and an XLSX (same base name).
        """
        if not product_list:
            print("[!] No data to save.")
            return

        # --- CSV ---
        with open(self.output_path, "w", newline="", encoding="utf-8-sig") as csvfile:
            # Dynamically grab fieldnames from the first item's keys
            fieldnames = list(product_list[0].keys())
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()
            for product in product_list:
                writer.writerow(product)

        print(f"[+] File CSV saved: {self.output_path}")

        # --- XLSX ---
        self.save_to_xlsx(product_list)

    def save_to_xlsx(self, product_list: list) -> None:
        """
        Saves the same data to an .xlsx file next to the CSV.
        Uses openpyxl directly to avoid extra dependencies at read time.
        """
        if not product_list:
            return  

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("[!] openpyxl not installed; skipping XLSX. Run: pip install openpyxl")
            return

        xlsx_path = os.path.splitext(self.output_path)[0] + ".xlsx"
        fieldnames = list(product_list[0].keys())

        wb = Workbook()
        ws = wb.active or wb.create_sheet()  # active sheet is always present on a new Workbook
        ws.title = "scraped_data"


        # Header row
        ws.append(fieldnames)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Data rows
        for product in product_list:
            ws.append([product.get(key, "") for key in fieldnames])

        # Formatting: wrap long text columns and set sensible widths
        wrap = Alignment(wrap_text=True, vertical="top")
        for i, name in enumerate(fieldnames, start=1):
            col_letter = get_column_letter(i)
            if name in ("description", "comments"):
                ws.column_dimensions[col_letter].width = 50
                for cell in ws[col_letter][1:]:  # skip header
                    cell.alignment = wrap
            else:
                ws.column_dimensions[col_letter].width = 15

        wb.save(xlsx_path)
        print(f"[+] File XLSX saved: {xlsx_path}")



    def run(self, max_posts=None):
        """
        The orchestrator method for the scraping lifecycle.
        """
        print(f"[*] Starting extraction for {self.target}...")
        if self.navigate_to_page():
            results = self.scrape_all_posts(max_posts=max_posts)
            self.save_to_csv(results)
            return results
        else:
            print(f"[!] Failed to navigate to {self.target}.")
            return []